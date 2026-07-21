from __future__ import annotations

import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
FIXTURE_DIR = REPO_ROOT / "tests" / "fixtures" / "tiny_rubrics_export"


def run_script(*args: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, *args],
        cwd=REPO_ROOT,
        text=True,
        capture_output=True,
        check=False,
    )


def test_extract_rubrics_to_workbook_writes_stacked_workbook(tmp_path: Path) -> None:
    output = tmp_path / "rubrics.xlsx"

    result = run_script(
        "scripts/extract_rubrics_to_workbook.py",
        str(FIXTURE_DIR),
        "--output",
        str(output),
    )

    assert result.returncode == 0, result.stderr
    assert output.exists()

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Rubrics", "Overall Levels"]

    rubrics_sheet = workbook["Rubrics"]
    overall_sheet = workbook["Overall Levels"]

    assert rubrics_sheet["A1"].value == "Sample Rubric One"
    assert rubrics_sheet["B3"].value == "Achieving\n90%+"
    assert rubrics_sheet["C3"].value == "Beginning\n50%+"
    assert rubrics_sheet["A4"].value == "Clarity"
    assert rubrics_sheet["B4"].value == "Fully explains the idea."
    assert rubrics_sheet["C4"].value == "Needs more detail."
    assert rubrics_sheet["D4"].value == 10
    assert rubrics_sheet["A6"].value == "Second Rubric / Two"

    assert overall_sheet["A2"].value == "Sample Rubric One"
    assert overall_sheet["B2"].value == "Achieving"
    assert overall_sheet["D2"].value == 90
    assert overall_sheet["A4"].value == "Second Rubric / Two"


def test_extract_rubrics_to_workbook_supports_one_sheet_per_rubric(tmp_path: Path) -> None:
    output = tmp_path / "rubrics-separate.xlsx"

    result = run_script(
        "scripts/extract_rubrics_to_workbook.py",
        str(FIXTURE_DIR),
        "--output",
        str(output),
        "--one-sheet-per-rubric",
    )

    assert result.returncode == 0, result.stderr
    assert output.exists()

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Sample Rubric One", "Second Rubric  Two", "Overall Levels"]
    assert workbook["Sample Rubric One"]["A1"].value == "Sample Rubric One"
    assert workbook["Second Rubric  Two"]["A1"].value == "Second Rubric / Two"


def test_extract_rubrics_to_workbook_creates_custom_output_parent(tmp_path: Path) -> None:
    output = tmp_path / "review" / "rubrics" / "rubrics.xlsx"

    result = run_script(
        "scripts/extract_rubrics_to_workbook.py",
        str(FIXTURE_DIR),
        "--output",
        str(output),
    )

    assert result.returncode == 0, result.stderr
    assert output.exists()


def test_extract_rubrics_json_output(tmp_path: Path) -> None:
    import json

    output = tmp_path / "rubrics.xlsx"
    json_output = tmp_path / "rubrics.json"

    result = run_script(
        "scripts/extract_rubrics_to_workbook.py",
        str(FIXTURE_DIR),
        "--output",
        str(output),
        "--json-output",
        str(json_output),
    )

    assert result.returncode == 0, result.stderr
    assert output.exists()
    assert json_output.exists()

    payload = json.loads(json_output.read_text(encoding="utf-8"))
    assert payload["schema"] == "coursecraft.rubrics/1"
    assert payload["source_file"] == "rubrics_d2l.xml"
    assert payload["rubrics"], "expected at least one rubric record"

    rubric = payload["rubrics"][0]
    for key in ("id", "resource_code", "name", "scoring_method", "levels", "criteria", "overall_levels", "attributes"):
        assert key in rubric, f"missing {key}"
    assert rubric["criteria"], "expected criteria rows"
    first_criterion = rubric["criteria"][0]
    assert first_criterion["cells"], "expected level cells"
    cell = first_criterion["cells"][0]
    for key in ("level_id", "level_name", "points", "points_raw", "description"):
        assert key in cell, f"missing cell key {key}"
    # levels carry the same names the workbook headers render
    level_names = {level["name"] for level in rubric["levels"]}
    assert cell["level_name"] in level_names or cell["level_name"] == ""
